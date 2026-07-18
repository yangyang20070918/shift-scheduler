from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
NOTE_FONT = Font(color="FF0000", size=9)
EXAMPLE_FONT = Font(color="808080", italic=True, size=9)


def generate_import_template(existing_data: dict | None = None) -> io.BytesIO:
    wb = Workbook()
    has_data = existing_data and existing_data.get("members")

    # --- 説明シート ---
    ws_guide = wb.active
    ws_guide.title = "説明"
    ws_guide.column_dimensions["A"].width = 60
    guide_lines = [
        "【インポート手順】",
        "1. まず画面上で「出勤パターン」を登録してください",
        "2. 「メンバー」シートにメンバー名と対応パターンを記入",
        "3. 「個人制約」シートに制約を記入（任意）",
        "4. 画面の「Excelインポート」からこのファイルをアップロード",
        "",
        "【注意事項】",
        "・パターン名は画面で登録済みの名前と完全一致が必要です",
        "・複数パターンはカンマ（,）区切りで記入",
        "・個人制約は未記入の項目は制約なしとして扱います",
        "・数値のみ入力してください（単位は不要）",
        "・既存メンバーの名前が一致する場合は更新、新しい名前は追加されます",
    ]
    for i, line in enumerate(guide_lines):
        cell = ws_guide.cell(row=i + 1, column=1, value=line)
        if line.startswith("【"):
            cell.font = Font(bold=True, size=11)
        else:
            cell.font = Font(size=10)

    # --- メンバーシート ---
    ws_members = wb.create_sheet("メンバー")
    headers = ["名前", "対応可能パターン（カンマ区切り）"]
    for ci, h in enumerate(headers):
        cell = ws_members.cell(row=1, column=ci + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    ws_members.column_dimensions["A"].width = 20
    ws_members.column_dimensions["B"].width = 40

    if has_data:
        for ri, m in enumerate(existing_data["members"], start=2):
            ws_members.cell(row=ri, column=1, value=m["name"])
            ws_members.cell(row=ri, column=2, value=",".join(m.get("pattern_names", [])))
    else:
        ws_members.cell(row=2, column=1, value="山田太郎").font = EXAMPLE_FONT
        ws_members.cell(row=2, column=2, value="日勤,夜勤").font = EXAMPLE_FONT
        ws_members.cell(row=3, column=1, value="佐藤花子").font = EXAMPLE_FONT
        ws_members.cell(row=3, column=2, value="日勤,夜勤,短時間").font = EXAMPLE_FONT

    # --- 個人制約シート ---
    ws_constraints = wb.create_sheet("個人制約")
    c_headers = [
        "メンバー名",
        "週出勤下限(日)", "週出勤上限(日)",
        "期間出勤下限(日)", "期間出勤上限(日)",
        "週労働下限(h)", "週労働上限(h)",
        "期間労働下限(h)", "期間労働上限(h)",
        "連続出勤上限(日)", "連休日数上限(日)",
    ]
    c_fields = [
        "weekly_work_days_min", "weekly_work_days_max",
        "period_work_days_min", "period_work_days_max",
        "weekly_work_hours_min", "weekly_work_hours_max",
        "period_work_hours_min", "period_work_hours_max",
        "max_consecutive_work_days", "max_consecutive_rest_days",
    ]
    for ci, h in enumerate(c_headers):
        cell = ws_constraints.cell(row=1, column=ci + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True)
        ws_constraints.column_dimensions[chr(65 + ci)].width = 14
    ws_constraints.column_dimensions["A"].width = 16

    if has_data and existing_data.get("constraints"):
        for ri, c in enumerate(existing_data["constraints"], start=2):
            ws_constraints.cell(row=ri, column=1, value=c.get("member_name", ""))
            for fi, field in enumerate(c_fields):
                val = c.get(field)
                if val is not None:
                    ws_constraints.cell(row=ri, column=fi + 2, value=val)
    else:
        ws_constraints.cell(row=2, column=1, value="山田太郎").font = EXAMPLE_FONT
        ws_constraints.cell(row=2, column=4, value=20).font = EXAMPLE_FONT
        ws_constraints.cell(row=2, column=5, value=23).font = EXAMPLE_FONT
        ws_constraints.cell(row=2, column=10, value=5).font = EXAMPLE_FONT

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


TYPE_FROM_LABEL = {"勤務": "work", "休日": "rest", "出張": "travel"}


def parse_pattern_excel(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    patterns: list[dict] = []
    sheet = None
    for name in ["出勤パターン", "Sheet1"]:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[0] if row else None
        if not name:
            continue
        p: dict = {"name": str(name).strip()}
        if len(row) > 1 and row[1]:
            label = str(row[1]).strip()
            p["type"] = TYPE_FROM_LABEL.get(label, label)
        else:
            p["type"] = "work"
        if len(row) > 2 and row[2]:
            p["start_time"] = str(row[2]).strip()
        if len(row) > 3 and row[3]:
            p["end_time"] = str(row[3]).strip()
        if len(row) > 4 and row[4] is not None:
            try:
                p["break_hours"] = float(row[4])
            except (ValueError, TypeError):
                pass
        if len(row) > 5 and row[5] is not None:
            try:
                p["work_hours"] = float(row[5])
            except (ValueError, TypeError):
                pass
        if len(row) > 6 and row[6]:
            p["color_code"] = str(row[6]).strip()
        patterns.append(p)

    wb.close()
    return patterns


def parse_schedule_result_excel(file_bytes: bytes) -> list[dict]:
    """Parse a schedule result Excel (排班表 sheet) and return assignment list."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    sheet = None
    for name in ["排班表", "Sheet1"]:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 4:
        return []

    header_row = rows[2]
    date_columns: list[tuple[int, str]] = []
    for ci, val in enumerate(header_row):
        if ci == 0 or val is None:
            continue
        date_columns.append((ci, str(val).strip()))

    assignments: list[dict] = []
    for row in rows[3:]:
        member_name = row[0] if row else None
        if not member_name:
            continue
        member_name = str(member_name).strip()
        for ci, date_label in date_columns:
            cell_val = row[ci] if ci < len(row) else None
            if cell_val is None or str(cell_val).strip() == "":
                continue
            val = str(cell_val).strip()
            is_rest = val == "休"
            assignments.append({
                "member_name": member_name,
                "date_label": date_label,
                "pattern_name": None if is_rest else val,
                "is_rest": is_rest,
            })

    return assignments


def parse_import_excel(file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    result: dict = {"members": [], "constraints": [], "demands": []}

    if "メンバー" in wb.sheetnames:
        ws = wb["メンバー"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            if not name:
                continue
            patterns_str = str(row[1] or "") if len(row) > 1 else ""
            pattern_names = [p.strip() for p in patterns_str.split(",") if p.strip()]
            result["members"].append({"name": str(name), "pattern_names": pattern_names})

    if "個人制約" in wb.sheetnames:
        ws = wb["個人制約"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            if not name:
                continue
            c: dict = {"member_name": str(name)}
            fields = [
                ("weekly_work_days_min", 1), ("weekly_work_days_max", 2),
                ("period_work_days_min", 3), ("period_work_days_max", 4),
                ("weekly_work_hours_min", 5), ("weekly_work_hours_max", 6),
                ("period_work_hours_min", 7), ("period_work_hours_max", 8),
                ("max_consecutive_work_days", 9), ("max_consecutive_rest_days", 10),
            ]
            for field, idx in fields:
                if len(row) > idx and row[idx] is not None:
                    try:
                        val = float(row[idx])
                        c[field] = val if "hours" in field else int(val)
                    except (ValueError, TypeError):
                        pass
            result["constraints"].append(c)

    if "毎日需要" in wb.sheetnames:
        ws = wb["毎日需要"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            date_val = row[0]
            if not date_val:
                continue
            d: dict = {"date": str(date_val)}
            if len(row) > 1 and row[1] is not None:
                try:
                    d["min_total"] = int(row[1])
                except (ValueError, TypeError):
                    pass
            if len(row) > 2 and row[2] is not None:
                try:
                    d["max_total"] = int(row[2])
                except (ValueError, TypeError):
                    pass
            result["demands"].append(d)

    wb.close()
    return result
