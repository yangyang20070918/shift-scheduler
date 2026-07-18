from __future__ import annotations

import io
import math
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TYPE_LABELS = {"work": "勤務", "rest": "休日", "travel": "出張"}
TYPE_FROM_LABEL = {v: k for k, v in TYPE_LABELS.items()}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
REST_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
WEEKEND_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
OK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NG_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
SUB_HEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SUB_HEADER_FONT = Font(bold=True, size=10)

WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]


def _styled_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=THIN_BORDER):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def generate_pattern_excel(patterns: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "出勤パターン"

    headers = ["名前", "タイプ", "開始時刻", "終了時刻", "休憩時間(h)", "実労働時間(h)", "色コード"]
    for ci, h in enumerate(headers):
        _styled_cell(ws, 1, ci + 1, h, font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="center"))
    col_widths = [16, 10, 12, 12, 14, 14, 12]
    for ci, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(ci + 1)].width = w

    for ri, p in enumerate(patterns, start=2):
        ws.cell(row=ri, column=1, value=p["name"])
        ws.cell(row=ri, column=2, value=TYPE_LABELS.get(p.get("type", "work"), p.get("type", "")))
        ws.cell(row=ri, column=3, value=p.get("start_time", ""))
        ws.cell(row=ri, column=4, value=p.get("end_time", ""))
        ws.cell(row=ri, column=5, value=p.get("break_hours", 0))
        ws.cell(row=ri, column=6, value=p.get("work_hours", 0))
        color = p.get("color_code", "#808080")
        ws.cell(row=ri, column=7, value=color)
        for ci in range(1, 8):
            ws.cell(row=ri, column=ci).border = THIN_BORDER
        if color and len(color.lstrip("#")) == 6:
            c = color.lstrip("#")
            ws.cell(row=ri, column=7).fill = PatternFill(start_color=c, end_color=c, fill_type="solid")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_schedule_excel(
    schedule_name: str,
    start_date: date,
    num_days: int,
    assignments: list[dict],
    violations: list[dict],
    warnings: list[dict],
    score_breakdown: dict | None,
    health_score: float | None,
    solve_time_seconds: float | None,
    members: list[dict],
    patterns: list[dict],
    daily_demands: list[dict] | None = None,
    pattern_demands: list[dict] | None = None,
    group_demands: list[dict] | None = None,
    groups: list[dict] | None = None,
) -> io.BytesIO:
    wb = Workbook()

    member_map = {m["id"]: m["name"] for m in members}
    pattern_map = {p["id"]: p for p in patterns}
    dates = [start_date + timedelta(days=i) for i in range(num_days)]

    _build_schedule_sheet(wb, schedule_name, dates, assignments, member_map, pattern_map)
    _build_daily_stats_sheet(wb, dates, assignments, member_map, pattern_map,
                             daily_demands or [], pattern_demands or [],
                             group_demands or [], groups or [])
    _build_personal_stats_sheet(wb, dates, assignments, member_map, pattern_map,
                                health_score, solve_time_seconds, score_breakdown)
    _build_violations_sheet(wb, violations, member_map)
    if warnings:
        _build_warnings_sheet(wb, warnings)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Sheet 1: 排班表 ──

def _build_schedule_sheet(wb, name, dates, assignments, member_map, pattern_map):
    ws = wb.active
    ws.title = "排班表"

    ws.cell(row=1, column=1, value=name).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(dates) + 1, 10))

    header_row = 3
    _styled_cell(ws, header_row, 1, "メンバー", font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="center"))
    ws.column_dimensions["A"].width = 14

    for di, d in enumerate(dates):
        col = di + 2
        wd = WEEKDAY_JP[d.weekday()]
        _styled_cell(ws, header_row, col, f"{d.month}/{d.day}({wd})",
                     font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="center"))
        ws.column_dimensions[get_column_letter(col)].width = 10

    member_ids = sorted(set(a["member_id"] for a in assignments))
    assignment_lookup: dict[tuple[str, str], dict] = {}
    for a in assignments:
        assignment_lookup[(a["member_id"], a["date"])] = a

    for mi, mid in enumerate(member_ids):
        row = header_row + 1 + mi
        name_cell = _styled_cell(ws, row, 1, member_map.get(mid, mid), font=Font(bold=True, size=10))

        for di, d in enumerate(dates):
            col = di + 2
            a = assignment_lookup.get((mid, str(d)))
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

            if d.weekday() >= 5:
                cell.fill = WEEKEND_FILL

            if a is None:
                cell.value = ""
            elif a.get("is_rest"):
                cell.value = "休"
                cell.fill = REST_FILL
                cell.font = Font(color="666666", size=10)
            else:
                pid = a.get("pattern_id")
                p = pattern_map.get(pid) if pid else None
                cell.value = p["name"] if p else (a.get("pattern_name") or "出勤")
                if p and p.get("color_code"):
                    color = p["color_code"].lstrip("#")
                    if len(color) == 6:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        r, g, b = int(color[:2], 16), int(color[2:4], 16), int(color[4:], 16)
                        text_color = "FFFFFF" if (r * 0.299 + g * 0.587 + b * 0.114) < 150 else "000000"
                        cell.font = Font(color=text_color, size=10)


# ── Sheet 2: 日別需要充足 ──

def _build_daily_stats_sheet(wb, dates, assignments, member_map, pattern_map,
                              daily_demands, pattern_demands, group_demands, groups):
    ws = wb.create_sheet("日別需要充足")
    ws.cell(row=1, column=1, value="日別需要充足").font = Font(bold=True, size=14)

    assignment_lookup: dict[str, list[dict]] = {}
    for a in assignments:
        assignment_lookup.setdefault(a["date"], []).append(a)

    row = 3

    # ── Section 1: 日別出勤統計 ──
    _styled_cell(ws, row, 1, "日別出勤統計", font=Font(bold=True, size=12))
    row += 1

    _styled_cell(ws, row, 1, "項目", font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="center"))
    ws.column_dimensions["A"].width = 16
    for di, d in enumerate(dates):
        col = di + 2
        wd = WEEKDAY_JP[d.weekday()]
        is_weekend = d.weekday() >= 5
        fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") if is_weekend else HEADER_FILL
        _styled_cell(ws, row, col, f"{d.month}/{d.day}({wd})",
                     font=HEADER_FONT, fill=fill, alignment=Alignment(horizontal="center"))
        ws.column_dimensions[get_column_letter(col)].width = 9

    demand_map = {d["date"]: d for d in daily_demands}

    for label, row_key in [("出勤人数", "actual"), ("必要人数", "required"), ("過不足", "diff")]:
        row += 1
        _styled_cell(ws, row, 1, label, font=SUB_HEADER_FONT, fill=SUB_HEADER_FILL)
        for di, d in enumerate(dates):
            col = di + 2
            ds = str(d)
            day_assignments = assignment_lookup.get(ds, [])
            actual = sum(1 for a in day_assignments if not a.get("is_rest"))
            demand = demand_map.get(ds)

            if row_key == "actual":
                _styled_cell(ws, row, col, actual, alignment=Alignment(horizontal="center"))
            elif row_key == "required":
                val = f"{demand['min_total']}~{demand['max_total']}" if demand else "-"
                _styled_cell(ws, row, col, val, alignment=Alignment(horizontal="center"))
            elif row_key == "diff":
                if demand:
                    diff = actual - demand["min_total"]
                    fill = OK_FILL if diff >= 0 else NG_FILL
                    val = f"+{diff}" if diff >= 0 else str(diff)
                    _styled_cell(ws, row, col, val, fill=fill, alignment=Alignment(horizontal="center"))
                else:
                    _styled_cell(ws, row, col, "-", alignment=Alignment(horizontal="center"))

    # ── Section 2: パターン別需要充足 ──
    if pattern_demands:
        row += 2
        _styled_cell(ws, row, 1, "パターン別需要充足", font=Font(bold=True, size=12))
        row += 1

        pd_by_pattern: dict[str, dict[str, int]] = {}
        for pd in pattern_demands:
            pd_by_pattern.setdefault(pd["pattern_id"], {})[pd["date"]] = pd["min_count"]

        _styled_cell(ws, row, 1, "パターン", font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="center"))
        for di, d in enumerate(dates):
            col = di + 2
            wd = WEEKDAY_JP[d.weekday()]
            is_weekend = d.weekday() >= 5
            fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") if is_weekend else HEADER_FILL
            _styled_cell(ws, row, col, f"{d.month}/{d.day}({wd})",
                         font=HEADER_FONT, fill=fill, alignment=Alignment(horizontal="center"))

        for pat_id, date_map in pd_by_pattern.items():
            row += 1
            p_name = pattern_map.get(pat_id, {}).get("name", pat_id)
            _styled_cell(ws, row, 1, p_name, font=SUB_HEADER_FONT, fill=SUB_HEADER_FILL)
            for di, d in enumerate(dates):
                col = di + 2
                ds = str(d)
                required = date_map.get(ds)
                if required is None:
                    _styled_cell(ws, row, col, "-", alignment=Alignment(horizontal="center"))
                else:
                    day_assignments = assignment_lookup.get(ds, [])
                    actual = sum(1 for a in day_assignments if a.get("pattern_id") == pat_id)
                    ok = actual >= required
                    fill = OK_FILL if ok else NG_FILL
                    font = Font(color="389E0D" if ok else "CF1322", bold=True, size=10)
                    _styled_cell(ws, row, col, f"{actual}/{required}", font=font, fill=fill,
                                 alignment=Alignment(horizontal="center"))

    # ── Section 3: グループ需要充足 ──
    if group_demands:
        row += 2
        _styled_cell(ws, row, 1, "グループ需要充足", font=Font(bold=True, size=12))
        row += 1

        group_map = {g["id"]: g for g in (groups or [])}

        combos: dict[str, dict] = {}
        for gd in group_demands:
            key = f"{gd['group_id']}_{gd['pattern_id'] or 'any'}"
            if key not in combos:
                g_name = group_map.get(gd["group_id"], {}).get("name", gd["group_id"])
                p_name = pattern_map.get(gd["pattern_id"], {}).get("name", "全パターン") if gd["pattern_id"] else "全パターン"
                combos[key] = {
                    "group_id": gd["group_id"],
                    "pattern_id": gd["pattern_id"],
                    "label": f"{g_name}({p_name})",
                    "date_map": {},
                }
            combos[key]["date_map"][gd["date"]] = gd["min_count"]

        _styled_cell(ws, row, 1, "グループ", font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="center"))
        for di, d in enumerate(dates):
            col = di + 2
            wd = WEEKDAY_JP[d.weekday()]
            is_weekend = d.weekday() >= 5
            fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") if is_weekend else HEADER_FILL
            _styled_cell(ws, row, col, f"{d.month}/{d.day}({wd})",
                         font=HEADER_FONT, fill=fill, alignment=Alignment(horizontal="center"))

        for combo in combos.values():
            row += 1
            _styled_cell(ws, row, 1, combo["label"], font=SUB_HEADER_FONT, fill=SUB_HEADER_FILL)
            group_members = group_map.get(combo["group_id"], {}).get("member_ids", [])
            for di, d in enumerate(dates):
                col = di + 2
                ds = str(d)
                required = combo["date_map"].get(ds)
                if required is None:
                    _styled_cell(ws, row, col, "-", alignment=Alignment(horizontal="center"))
                else:
                    day_assignments = assignment_lookup.get(ds, [])
                    working = [a for a in day_assignments if a["member_id"] in group_members and not a.get("is_rest")]
                    if combo["pattern_id"]:
                        actual = sum(1 for a in working if a.get("pattern_id") == combo["pattern_id"])
                    else:
                        actual = len(working)
                    ok = actual >= required
                    fill = OK_FILL if ok else NG_FILL
                    font = Font(color="389E0D" if ok else "CF1322", bold=True, size=10)
                    _styled_cell(ws, row, col, f"{actual}/{required}", font=font, fill=fill,
                                 alignment=Alignment(horizontal="center"))


# ── Sheet 3: 個人統計 + 全体バランス ──

def _build_personal_stats_sheet(wb, dates, assignments, member_map, pattern_map,
                                 health_score, solve_time, score_breakdown):
    ws = wb.create_sheet("統計情報")
    ws.cell(row=1, column=1, value="統計情報").font = Font(bold=True, size=14)

    row = 3
    if health_score is not None:
        ws.cell(row=row, column=1, value="健全スコア").font = Font(bold=True)
        ws.cell(row=row, column=2, value=round(health_score, 1))
        row += 1
    if solve_time is not None:
        ws.cell(row=row, column=1, value="求解時間").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{solve_time:.2f}秒")
        row += 1

    if score_breakdown:
        row += 1
        ws.cell(row=row, column=1, value="スコア明細").font = Font(bold=True, size=12)
        row += 1
        label_map = {"personal": "個人制約", "group": "グループ需要", "demand": "毎日需要", "balance": "均衡性", "total_penalty": "合計ペナルティ"}
        for key, label in label_map.items():
            if key in score_breakdown:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=score_breakdown[key])
                row += 1

    # ── Per-member stats with pattern breakdown ──
    row += 1
    ws.cell(row=row, column=1, value="メンバー別統計").font = Font(bold=True, size=12)
    row += 1

    work_pattern_ids = [pid for pid, p in pattern_map.items() if p.get("type") in ("work", "travel", "NORMAL", "TRAVEL")]
    base_headers = ["メンバー", "出勤日数", "休息日数", "合計労働時間"]
    pattern_headers = [pattern_map[pid]["name"] for pid in work_pattern_ids]
    all_headers = base_headers + pattern_headers

    for ci, h in enumerate(all_headers):
        _styled_cell(ws, row, ci + 1, h, font=HEADER_FONT, fill=HEADER_FILL)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    for ci in range(len(pattern_headers)):
        ws.column_dimensions[get_column_letter(5 + ci)].width = 10

    member_ids = sorted(set(a["member_id"] for a in assignments))
    all_work_days = []
    all_total_hours = []

    for mid in member_ids:
        row += 1
        member_assignments = [a for a in assignments if a["member_id"] == mid]
        work_days = sum(1 for a in member_assignments if not a.get("is_rest"))
        rest_days = sum(1 for a in member_assignments if a.get("is_rest"))
        total_hours = 0.0
        pattern_counts: dict[str, int] = {}
        for a in member_assignments:
            if not a.get("is_rest") and a.get("pattern_id"):
                p = pattern_map.get(a["pattern_id"])
                if p:
                    total_hours += p.get("work_hours", 0)
                    pattern_counts[a["pattern_id"]] = pattern_counts.get(a["pattern_id"], 0) + 1

        all_work_days.append(work_days)
        all_total_hours.append(total_hours)

        _styled_cell(ws, row, 1, member_map.get(mid, mid))
        _styled_cell(ws, row, 2, work_days)
        _styled_cell(ws, row, 3, rest_days)
        _styled_cell(ws, row, 4, round(total_hours, 1))
        for pi, pid in enumerate(work_pattern_ids):
            _styled_cell(ws, row, 5 + pi, pattern_counts.get(pid, 0))

    # ── Overall balance ──
    if all_work_days:
        row += 2
        ws.cell(row=row, column=1, value="全体バランス").font = Font(bold=True, size=12)
        row += 1

        avg_wd = sum(all_work_days) / len(all_work_days)
        avg_h = sum(all_total_hours) / len(all_total_hours)
        stddev_wd = math.sqrt(sum((x - avg_wd) ** 2 for x in all_work_days) / len(all_work_days))

        balance_data = [
            ("出勤日数（平均）", f"{avg_wd:.1f}日"),
            ("出勤日数（最小/最大）", f"{min(all_work_days)} / {max(all_work_days)}日"),
            ("出勤日数（標準偏差）", f"{stddev_wd:.2f}日"),
            ("労働時間（平均）", f"{avg_h:.1f}h"),
            ("労働時間（最小/最大）", f"{min(all_total_hours):.1f} / {max(all_total_hours):.1f}h"),
        ]
        for label, val in balance_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=val)
            row += 1


# ── Sheet 4: 違反レポート ──

def _build_violations_sheet(wb, violations, member_map):
    ws = wb.create_sheet("違反レポート")
    ws.cell(row=1, column=1, value="違反レポート").font = Font(bold=True, size=14)

    if not violations:
        ws.cell(row=3, column=1, value="違反はありません").font = Font(color="008000", size=12)
        return

    headers = ["グループ", "制約種別", "優先度", "メンバー", "対象日", "設定値", "実績値", "原因分析", "改善提案"]
    col_widths = [12, 20, 8, 14, 12, 10, 10, 30, 30]
    row = 3
    for ci, h in enumerate(headers):
        _styled_cell(ws, row, ci + 1, h, font=HEADER_FONT, fill=HEADER_FILL)
        ws.column_dimensions[get_column_letter(ci + 1)].width = col_widths[ci]

    constraint_labels = {
        "daily_demand_min": "毎日需要（最小）",
        "daily_demand_max": "毎日需要（最大）",
        "period_days_min": "期間出勤日数（最小）",
        "period_days_max": "期間出勤日数（最大）",
        "weekly_days_min": "週出勤日数（最小）",
        "weekly_days_max": "週出勤日数（最大）",
        "consecutive_work": "連続出勤超過",
        "consecutive_rest": "連続休息超過",
        "group_demand_min": "グループ需要（最小）",
        "period_hours_min": "期間労働時間（最小）",
        "period_hours_max": "期間労働時間（最大）",
    }
    group_labels = {"personal": "個人制約", "demand": "需要制約", "group": "グループ制約"}

    for v in violations:
        row += 1
        _styled_cell(ws, row, 1, group_labels.get(v.get("constraint_group", ""), v.get("constraint_group", "")))
        _styled_cell(ws, row, 2, constraint_labels.get(v.get("constraint_type", ""), v.get("constraint_type", "")))
        _styled_cell(ws, row, 3, v.get("priority", ""))
        mid = v.get("target_member_id")
        _styled_cell(ws, row, 4, member_map.get(mid, mid or ""))
        _styled_cell(ws, row, 5, v.get("target_date", ""))
        _styled_cell(ws, row, 6, v.get("setting_value", ""))
        _styled_cell(ws, row, 7, v.get("actual_value", ""))
        factors = v.get("contributing_factors", [])
        cell = _styled_cell(ws, row, 8, "\n".join(factors) if factors else "")
        cell.alignment = Alignment(wrap_text=True)
        suggestions = v.get("suggestions", [])
        cell = _styled_cell(ws, row, 9, "\n".join(suggestions) if suggestions else "")
        cell.alignment = Alignment(wrap_text=True)


# ── Sheet 5: 警告 ──

def _build_warnings_sheet(wb, warnings):
    ws = wb.create_sheet("警告")
    ws.cell(row=1, column=1, value="警告一覧").font = Font(bold=True, size=14)

    headers = ["種別", "重要度", "メッセージ"]
    row = 3
    for ci, h in enumerate(headers):
        _styled_cell(ws, row, ci + 1, h, font=HEADER_FONT, fill=HEADER_FILL)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 50

    for w in warnings:
        row += 1
        _styled_cell(ws, row, 1, w.get("warning_type", ""))
        _styled_cell(ws, row, 2, w.get("severity", ""))
        _styled_cell(ws, row, 3, w.get("message", ""))
